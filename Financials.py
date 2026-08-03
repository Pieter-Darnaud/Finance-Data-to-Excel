# Import yfinance tool:

import yfinance as yf


import openpyxl as opx



# To test if companies have a value, use this: 
def get_financials(companyTicker):
    
    company = yf.Ticker(companyTicker).financials
    try:
        return (
            "Ticker: "   + companyTicker
             + f", total revenue: ${company.loc["Total Revenue"].iloc[0]:.2f}"
            + f", cost of revenue: ${company.loc["Cost Of Revenue"].iloc[0]:.2f}"
            
        )
    except KeyError as e:
        print(f"  skipping {companyTicker}: no {e} line")
    
        return None 

# Testing try-catch operation
get_financials("AAPL")

 
# Standalone accounting functions

def cogs(bInventory, purchases, eInventory):
    return bInventory + purchases - eInventory

def grossProfit(revenue, cogs):
    return revenue - cogs

def opIncome(grossProfit, opex):
    return grossProfit - opex

def grossMargin(revenue, cogs):
    if cogs == 0:
        return "N/A"
    
    return (revenue - cogs) / revenue

def netMargin (netIncome, revenue):
    return netIncome / revenue

def opMargin (opIncome, revenue):
    return opIncome / revenue

# Define Apple, Microsoft and Bank of America by Ticker 
t = yf.Ticker("MSFT")
   

b = yf.Ticker("BAC")

a = yf.Ticker("AAPL")




 

# Establish variables for Microsoft and Bank of America
Mrevenue = t.financials.loc["Total Revenue"].iloc[0]
Mcogs    = t.financials.loc["Cost Of Revenue"].iloc[0]




Brevenue = b.financials.loc["Total Revenue"].iloc[0]
Bcogs    = 0


#Definitions and classing of various companies
Microsoft = { "ticker" : "MSFT" , "revenue" : Mrevenue, "cogs" : Mcogs}
BankOfAmerica = { "ticker" : "BAC" , "revenue" : Brevenue, "cogs" : Bcogs}
companies = [Microsoft, BankOfAmerica]

#Establish Apple's data and addition to companies

Arevenue = a.financials.loc["Total Revenue"].iloc[0]
Acogs    = a.financials.loc["Cost Of Revenue"].iloc[0]


Apple = {"ticker": "AAPL", "revenue": Arevenue, "cogs": Acogs}

companies.append(Apple)

# Testing out gross margin function
print ("Bank of America's Gross Margin: " + grossMargin(BankOfAmerica["revenue"], BankOfAmerica["cogs"]))

# Formatting each company function
for i in (companies):
    flag = ""

    if grossMargin(i["revenue"], i["cogs"]) == "N/A":
        flag = "no gross margin"
    elif grossMargin(i["revenue"], i["cogs"]) < 0.4:
        flag = "low"
    elif 0.4 <= grossMargin(i["revenue"], i["cogs"]) < 0.7:
        flag = "medium"
    else:
        flag = "high"

    if grossMargin(i["revenue"], i["cogs"]) == "N/A":
        print( i["ticker"] + ": " + flag)
    else:
        print(f"{i["ticker"]}: {grossMargin(i["revenue"], i["cogs"]):.2%} " + flag  )

# Excel sheet createion
w = opx.Workbook()
ws = w.active


# Sheet establishment and formatting
ws["A1"] = "Ticker"

ws["B1"] = "Revenue ($)"

ws["C1"] = "Cost of Goods Sold ($)"

ws["D1"] = "Gross Margin"

ws["E1"] = "Flag"

# This loop returns ticker, revenue, cogs, and gross margin values
for i in range (len(companies)):
    flag = ""
    if grossMargin(companies[i]["revenue"], companies[i]["cogs"]) == "N/A":
            flag = "no gross margin"
    elif grossMargin(companies[i]["revenue"], companies[i]["cogs"]) < 0.4:
            flag = "low"
    elif 0.4 <= grossMargin(companies[i]["revenue"], companies[i]["cogs"]) < 0.7:
            flag = "medium"
    else:
            flag = "high"
    ws["A" + str((i + 2))] = companies[i]["ticker"]
    ws["B" + str((i + 2))] = companies[i]["revenue"]
    ws["B" + str((i + 2))].number_format = "#,##0"
    ws["C" + str((i + 2))] = companies[i]["cogs"]
    ws["C" + str((i + 2))].number_format = "#,##0"
    ws["D" + str((i + 2))] = grossMargin(companies[i]["revenue"], companies[i]["cogs"])
    ws["D" + str((i + 2))].number_format = "0.0%"
    ws["E" + str((i + 2))] = flag



# Save a new excel sheet when this runs
w.save("FinancialsPuller.xlsx")
    


