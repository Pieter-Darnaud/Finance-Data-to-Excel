import openpyxl as opx

from metrics import grossMargin

from data import companies
low = 0.4

medium = 0.7


# Excel sheet createion
w = opx.Workbook()
ws = w.active


# Sheet establishment and formatting
ws["A1"] = "Ticker"

ws["B1"] = "Revenue ($)"

ws["C1"] = "Cost of Goods Sold ($)"

ws["D1"] = "Gross Margin"

ws["E1"] = "Flag"

# This loop returns ticker, revenue, cogs, and gross margin values
for i in range (len(companies)):
    flag = ""
    if grossMargin(companies[i]["revenue"], companies[i]["cogs"]) == "N/A":
            flag = "no gross margin"
    elif grossMargin(companies[i]["revenue"], companies[i]["cogs"]) < low:
            flag = "low"
    elif low <= grossMargin(companies[i]["revenue"], companies[i]["cogs"]) < medium:
            flag = "medium"
    else:
            flag = "high"
    ws["A" + str((i + 2))] = companies[i]["ticker"]
    ws["B" + str((i + 2))] = companies[i]["revenue"]
    ws["B" + str((i + 2))].number_format = "#,##0"
    ws["C" + str((i + 2))] = companies[i]["cogs"]
    ws["C" + str((i + 2))].number_format = "#,##0"
    ws["D" + str((i + 2))] = grossMargin(companies[i]["revenue"], companies[i]["cogs"])
    ws["D" + str((i + 2))].number_format = "0.0%"
    ws["E" + str((i + 2))] = flag



# Save a new excel sheet when this runs
w.save("FinancialsPuller.xlsx")
    


